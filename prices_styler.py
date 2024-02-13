from openpyxl.styles import PatternFill
from openpyxl import load_workbook

class ExcelColorChanger:
    def __init__(self, file_path):
        self.file_path = file_path
        self.frozen_pane_height = 0  # Initialize the attribute

    def change_cell_color(self):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            frozen_pane = sheet.sheet_view.pane
            self.frozen_pane_height = int(frozen_pane.ySplit) if frozen_pane.ySplit is not None else 0  # Store the value in the instance attribute
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            for row in sheet.iter_rows(min_row=self.frozen_pane_height + 1, max_row=sheet.max_row, min_col=1):
                min_value = float('inf')
                min_column = None

                for cell in row:
                    if cell.value is not None and isinstance(cell.value, (float, int)):
                        if cell.value < min_value:
                            min_value = cell.value
                            min_column = cell.column

                if min_column:
                    sheet.cell(row=row[0].row, column=min_column).fill = light_green_fill
            wb.save(self.file_path)
            wb.close()
        except Exception as e:
            print(f"An error occurred with coloring the cells: {e}")

# After processing and updating the Excel sheet with product information
file_path = "C:/Users/Slambe/Desktop/UNI/ohjelmointi/Python/SOK/kilpailijahintoja.xlsx"
color_changer = ExcelColorChanger(file_path)
color_changer.change_cell_color()
