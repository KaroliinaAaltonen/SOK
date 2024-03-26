import time
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import concurrent.futures
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils import get_column_letter
import openpyxl.styles
import winsound
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
# web scraping functions
# driver is started locally (and ended)
# get_html_from_url() uses sleep=10 with puuilo.fi and tokmanni.fi as per their robots.txt
class Scraper:
    def __init__(self):
        self.driver = self.local_driver()

    def local_driver(self):
        chrome_options = Options()
        chrome_options.add_argument('--headless')  # Run Chrome in headless mode (without UI)
        service = ChromeService(executable_path=ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        return driver

    def end_session(self):
        self.driver.quit()

    def get_html_from_url(self, url, sleep=1):
        try:
            self.driver.get(url)
            time.sleep(sleep)
            return self.driver.page_source
        except Exception as e:
            print(f"Error getting HTML from {url}: {str(e)}")
            return None

class PrismaScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.prisma.fi/haku?search={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    raise ValueError("\nPRISMA: Tuotesivua ei löytynyt")
                link = self.specific_product_page(soup)
                link = f"https://www.prisma.fi{link}"
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    product_name, brand_name, price = self.extract_product_information(soup, product_code)
                    if product_name and brand_name:
                        return product_name, link, brand_name, price
                else:
                    raise ValueError("(1) PRISMA: Error with handling HTML.")
            else:
                raise ValueError("(2) PRISMA: Error with handling HTML.")
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): #returns False if there were no search results or True if there were results
        try:
            no_search_results_element = soup.find('p', {'data-test-id': 'no-search-results', 'class': 'search_noResultTitle__PTo4Q'})
            if no_search_results_element:
                result_text = no_search_results_element.get_text(strip=True)
                return False
            else:
                return True
        except Exception as e:
            return None

    def specific_product_page(self, soup):
        try:
            link_tag = soup.find('a', class_='ProductCard_blockLink__NTUGB') # Find the <a> tag with the specified class
            if link_tag:
                link = link_tag['href'] # Extract the value of the href attribute
                return link
            else:
                return None
        except Exception as e:
            return None
        
    def extract_product_information(self, soup, product_code):
        try:
            # Find the <div> element...
            div_elements = soup.find_all('div', class_='Accordion_accordionContent__rtLYZ accordion-content')
            # ...that contains the div...
            for div in div_elements:
                # ...that contains the <dt>-tags...
                dt_tags = div.find_all('dt')
                for dt_tag in dt_tags:
                    # ...out of which one contains the product EAN
                    if dt_tag.text.strip() == 'Tuotekoodi':
                        dd_tag = dt_tag.find_next_sibling('dd')
                        ean_value = dd_tag.text.strip()
                        if(int(ean_value) != product_code):
                            return ValueError("EAN code does not match prisma.fi product code")
                        break
            # Find the element with class 'ProductMainInfo_brand__NbMHp'
            brand_div = soup.find('div', class_='ProductMainInfo_brand__NbMHp')
            # Extract the brand text from the <a> tag inside the div
            brand_name = brand_div.find('a').text

            # Find the element with class 'ProductMainInfo_finalPrice__1hFhA'
            price_span = soup.find('span', class_='ProductMainInfo_finalPrice__1hFhA')
            # Extract the price text from the span and remove non-numeric characters
            price_text = price_span.text.replace('&nbsp;', '').replace('€', '').replace(',', '.')
            # Convert the extracted text to a float
            price = float(price_text)

            # Find the element with class 'ProductMainInfo_title__VKU68' and data-test-id 'product-name'
            product_name_h1 = soup.find('h1', class_='ProductMainInfo_title__VKU68', attrs={'data-test-id': 'product-name'})
            # Extract the text content from the h1 element
            product_name = product_name_h1.text
            
            return product_name, brand_name, price
        except Exception as e:
            return None, None, None

class PuuiloScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.puuilo.fi/catalogsearch/result/?q={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                self.search_result_check(soup)                
                link = self.specific_product_page(soup)
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup, product_code)
                    if product_name and price:
                        return link, price, product_name
                else:
                    raise ValueError("(1) PUUILO: Error with handling HTML.")
            else:
                raise ValueError("(2) PUUILO: Error with handling HTML.")
        except Exception as e:
            return None, None, None
    
    def search_result_check(self, soup): # returns True if there are search results or False if no results
        hits_div = soup.find('div', class_='hits')
        result_count = hits_div.find('span', class_='ais-Stats-text')
        if result_count and result_count.get_text(strip=True) == '0 hakutulosta':
            return False
        else:
            return True
        
    def specific_product_page(self, soup): # returns link to the prodcut page or None
        try:
            div_list_content = soup.find('div', class_='list-content')
            div_result_wrapper = div_list_content.find('div', class_='result-wrapper')
            a_element = div_result_wrapper.find('a', class_='result') # Find the <a> element inside the parent div tag
            if a_element:
                link = a_element['href'] # Extract the value of the "href" attribute
                return link
            else:
                return ValueError("PUUILO: Tuotesivua ei löytynyt.")
        except Exception as e:
            return None

    def extract_product_information(self, soup, product_code): # returns price and product name or None
        try:
            # Find the <div> where EAN code is nested
            ean_parent_div = soup.find('div', class_='product-info-main')
            ean_child_div = ean_parent_div.find('div', class_='product-sku-container')
            ean_content_div = ean_child_div.find('div', class_='product attribute ean')
            ean = ean_content_div.find('div', class_='value').get_text()
            if(int(ean) != product_code):
                return ValueError("PUUILO: EAN-koodi ei vastannut hakua")
            # Find the element with class 'price' inside the span with id 'product-price-107111'
            price_span = soup.find('span', class_='price')
            # Extract the price text from the span and remove non-numeric characters
            price_text = price_span.text.replace('&nbsp;', '').replace('€', '').replace(',', '.')
            # Convert the extracted text to a float
            price = float(price_text)

            # Extract product name
            product_name = soup.find('span', {'data-ui-id': 'page-title-wrapper'}).text.strip()

            return price, product_name
        except Exception as e:
            return None, None

class KRautaScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.k-rauta.fi/etsi?query={product_code}"
            html = self.get_html_from_url(url)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                if not self.search_result_check(soup):
                    raise ValueError("\nK-RAUTA:\nEi tuloksia haulla:", product_code)
                
                link = self.specific_product_page(soup)
                link = f"https://www.k-rauta.fi{link}"
                html = self.get_html_from_url(link)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup)
                    if product_name and price:
                        return link, price, product_name
                else:
                    raise ValueError("(1) K-RAUTA: Error with handling HTML.")
            else:
                raise ValueError("(2) K-RAUTA: Error with handling HTML.")
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): # returns True if there are search results or False if no results
        try:
            no_search_results_element = soup.find('div', class_="empty-search-result")
            if no_search_results_element:
                result_text = no_search_results_element.get_text(strip=True)
                return False
            else:
                return True
        except Exception as e:
            return None
        
    def specific_product_page(self, soup): # returns specific product page or None
        try:
            a_element = soup.find('a', class_='product-card') # Find the <a> element with class "product-card"
            if a_element:
                link = a_element['href'] # Extract the value of the "href" attribute
                return link
            else:
                return None
        except Exception as e:
            return None

    def extract_product_information(self, soup): # returns price and product name or None
        try:
            # Extract price span
            price_span = soup.find('span', class_='price-view-fi__sale-price--prefix')
            price_text = price_span.next_sibling.strip()
            price= float(price_text.replace("€", "").replace(",", "."))

            # Extract the product name
            product_name = soup.find('h1', class_='product-heading__product-name').text.strip()
            return price, product_name
        except Exception as e:
            return None, None

class TokmanniScraper(Scraper):
    def __init__(self):
        super().__init__()
    def search_product(self, product_code):
        try:
            url = f"https://www.tokmanni.fi/search/?q={product_code}"
            html = self.get_html_from_url(url, 10)
            if html:
                soup = BeautifulSoup(html, 'html.parser')
                self.search_result_check(soup)
                link = self.specific_product_page(soup, product_code)
                html = self.get_html_from_url(link, 10)
                if html:
                    soup = BeautifulSoup(html, 'html.parser')
                    price, product_name = self.extract_product_information(soup, product_code)
                    if product_name and price:
                        return link, price, product_name
                else:
                    raise ValueError("(1) TOKMANNI: Error with handling HTML.")
            else:
                raise ValueError("(2) TOKMANNI: Error with handling HTML.")
        except Exception as e:
            return None, None, None

    def search_result_check(self, soup): # returns False if no results found, True if results were found, otherwise None
        try:
            # Define the target text
            target_text = "Kokeile toista hakusanaa ..."
            # Find the wanted HTML
            page_wrapper = soup.find('div', class_='page-wrapper')
            page_columns = page_wrapper.find('div', class_='columns')
            main_column = page_columns.find('div', class_='column main')
            no_results_parent = main_column.find('div', class_='kuContainer kuFiltersLeft')
            no_results = no_results_parent.find('div', class_='kuNoRecordFound')
            child_1 = no_results.find('div', class_='kuNoResults-lp')
            child_2 = child_1.find('div', class_='kuNoResultsInner')
            target_div = child_2.find('div', class_='kuNoResults-lp-message')
            no_results_text = target_div.get_text()
            # Check if the target text is present in the div
            if target_text in no_results_text:
                return ValueError("TOKMANNI: tuotetta ei löytynyt hausta")
            else:
                return True
        except Exception as e:
            return ValueError("TOKMANNI: tuotetta ei löytynyt hausta")

    def specific_product_page(self, soup, product_code): # Returns the link to the product page ro None
        try:
            # Find the wanted HTML
            page_wrapper = soup.find('div', class_='page-wrapper')
            page_columns = page_wrapper.find('div', class_='columns')
            main_column = page_columns.find('div', class_='column main')
            results_parent = main_column.find('div', class_='kuContainer kuFiltersLeft')
            results = results_parent.find('div', class_='kuProListing')
            child_1 = results.find('div', class_='kuResultList')
            child_2 = child_1.find('div', class_='kuProductContent')
            a_tag = child_2.find('div', class_='kuGridView').find('a')
            link = a_tag['href']
            # Check if the link was found
            if link:
                return link
            else:
                return ValueError("TOKMANNI: tuotesivua ei löytynyt.")
        except Exception as e:
            return ValueError("TOKMANNI: tuotesivua ei löytynyt.")
        
    def extract_product_information(self, soup, product_code): # Returns price and product name or None
        try:
            ean_div = soup.find('div', class_='product attribute sku')
            ean_value = ean_div.find('div', class_='value').get_text(strip=True)
            if(str(ean_value) != str(product_code)):
                return ValueError("Tokamnni: Tuotesivun EAN ei täsmää haettua tuotetta.")
            # Find the span element with class "price-wrapper"
            price_wrapper_span = soup.find('span', class_='price-wrapper')
            # Extract the value from the data-price-amount attribute
            price = price_wrapper_span.get('data-price-amount')
            price = float(price)
            # Find the h1 element with class "page-title"
            page_title_h1 = soup.find('h1', {'class': 'page-title'})
            # Extract product name
            product_name = page_title_h1.get_text(strip=True)
            return price, product_name
        except Exception as e:
            return ValueError("Tokmanni: Virhe tuotetietojen hakemisessa.")
            
def format_product_link(url, product_name, sheet, given_row, given_column):
    try:
        cell = sheet.cell(row=given_row, column=given_column)
        cell.value = product_name
        cell.hyperlink = url
        return None
    except Exception as e:
        return None

def format_EAN(product_code):
    try:
        # Convert product_code to an integer
        product_code = int(product_code)
        # Raise ValueError if the input does not match EAN-13 length
        if len(str(product_code)) != 13:
            raise ValueError("Invalid EAN:", product_code, "Input string must have exactly 13 characters.")
        return product_code
    except Exception as e:
        return None
    
class ExcelHandler:
    def handle(self, file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            frozen_pane = sheet.sheet_view.pane
            frozen_pane_height = int(frozen_pane.ySplit) if frozen_pane.ySplit is not None else 0
            target_column_index = 1
            blank_cell_encountered = False

            with concurrent.futures.ThreadPoolExecutor() as executor:
                # Create instances of scraper classes directly
                prisma_scraper = PrismaScraper()
                k_rauta_scraper = KRautaScraper()
                puuilo_scraper = PuuiloScraper()
                tokmanni_scraper = TokmanniScraper()
                # List of functions to execute in parallel
                functions_to_execute = [
                    prisma_scraper.search_product,
                    k_rauta_scraper.search_product,
                    puuilo_scraper.search_product,
                    tokmanni_scraper.search_product,
                ]

                for row in sheet.iter_rows(min_row=frozen_pane_height + 1, max_row=sheet.max_row, min_col=1, max_col=12):
                    if target_column_index >= len(row):
                        new_column = sheet.max_column + 1
                        sheet.cell(row=1, column=new_column, value=f'Column{new_column}')

                    if row[0].value is not None:
                        # Execute functions in parallel
                        product_code = format_EAN(row[0].value)
                        results = list(executor.map(lambda f: f(product_code), functions_to_execute))
                        # Process results and update the sheet accordingly
                        if all(result is not None for result in results[0][:4]):  # PRISMA INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 1, value=results[0][0].upper())  # product name
                            sheet.cell(row=row[0].row, column=target_column_index + 2, value=results[0][2])  # brand
                            sheet.cell(row=row[0].row, column=target_column_index + 3, value=results[0][3])  # price
                            current_row = row[0].row
                            current_column = target_column_index + 1
                            format_product_link(results[0][1], results[0][0], sheet, current_row, current_column)  # link with product name as label
                        if results[1]:  # K-RAUTA INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 4, value=results[1][1])  # price
                            current_row = row[0].row
                            current_column = target_column_index + 5
                            format_product_link(results[1][0], results[1][2], sheet, current_row, current_column)  # link with product name as label
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 4, value="")  # leave K-rauta price empty
                            sheet.cell(row=row[0].row, column=target_column_index + 5, value="")  # leave K-rauta link empty
                        if results[2]:  # PUUILO INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 6, value=results[2][1])
                            current_row = row[0].row
                            current_column = target_column_index + 7
                            format_product_link(results[2][0], results[2][2], sheet, current_row, current_column)
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 6, value="")
                            sheet.cell(row=row[0].row, column=target_column_index + 7, value="")

                        if results[3]:  # TOKMANNI INFO TO EXCEL
                            sheet.cell(row=row[0].row, column=target_column_index + 8, value=results[3][1])
                            current_row = row[0].row
                            current_column = target_column_index + 9
                            format_product_link(results[3][0], results[3][2], sheet, current_row, current_column)
                        else:
                            sheet.cell(row=row[0].row, column=target_column_index + 8, value="")
                            sheet.cell(row=row[0].row, column=target_column_index + 9, value="")
                    else:
                        blank_cell_encountered = True

                    if blank_cell_encountered:
                        break

        except Exception as e:
            raise ValueError(f"An error occurred with Excel handle function: {str(e)}")
        finally:
            try:
                if 'wb' in locals() and wb is not None:
                    wb.save(file_path)
                    wb.close()
            except Exception as e:
                print(f"An error occurred while saving/closing the workbook: {e}")


class ExcelColorChanger:
    def __init__(self, file_path):
        self.file_path = file_path
        self.frozen_pane_height = 0  # Initialize the attribute

    def change_cell_color(self):
        try:
            wb = load_workbook(self.file_path)
            sheet = wb.active
            frozen_pane = sheet.sheet_view.pane
            self.frozen_pane_height = int(frozen_pane.ySplit) if frozen_pane.ySplit is not None else 0  # Store the value in the instance attribute
            light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            for row in sheet.iter_rows(min_row=self.frozen_pane_height + 1, max_row=sheet.max_row, min_col=1):
                min_value = float('inf')
                min_column = None

                if row[3].value is not None:  # Check if the fourth cell is not empty
                    for cell in row:  # Iterate over cells in the row
                        if cell.value is not None and isinstance(cell.value, (float, int)):
                            if cell.value < min_value:
                                min_value = cell.value
                                min_column = cell.column

                if min_column:
                    sheet.cell(row=row[0].row, column=min_column).fill = light_green_fill
            wb.save(self.file_path)
            wb.close()
        except Exception as e:
            raise ValueError(f"An error occurred with coloring the cells: {e}")

# Main
##################################################################################################
# JOS SÄ ET OSAA OHJELMOIDA NIIN ÄLÄ KOSKE MIHINKÄÄN PAITSI TÄHÄN JOS ON IHAN PAKKO              #
# LAITA SUN EAN-KOODIT SISÄLTÄVÄN EXCEL-TIEDOSTON SIJAINTI TON ALLA OLEVAN TEKSTIN               #
# "file_path =" perään HEITTOMERKKIEN SISÄÄN esimerkiksi JOS SUN TIEDOSOTO ON                    #
# SUN TIETOKONEELLA SIJAINNISSA "C:/Users/Karoliina/kilpailijahintoja.xlsx" NIIN                 #
# MUUTA TOI RIVI NIIN ETTÄ SIINÄ LUKEE: file_path = "C:/Users/Karoliina/kilpailijahintoja.xlsx"  #
##################################################################################################
file_path = "C:/Users/Karoliina/Desktop/kilpailijahintoja.xlsx"

start_time = time.time()
handler = ExcelHandler()
handler.handle(file_path)
color_changer = ExcelColorChanger(file_path)
color_changer.change_cell_color()
end_time = time.time()
elapsed_time = end_time - start_time
# Play a sound when done
if __name__ == "__main__":
    print(f"This program took: {elapsed_time:.2f} seconds to run. Kiitos ja anteeksi.")
    for _ in range(3):
        winsound.Beep(1000, 500)
